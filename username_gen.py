import openpyxl
import random
import argparse

parser = argparse.ArgumentParser()

parser.add_argument("--file", "-f", type=str, required=True)
args = parser.parse_args()

# load excel with its path
Workbook = openpyxl.load_workbook(args.file)

#Defining a specific sheet if needed
# sh = Workbook['Sheet2']


def edit_excel():
    print("Reading the input file")
    for sheet in Workbook.worksheets:
        sh = Workbook[sheet.title]
        print("Processing {}".format(sheet.title))
        for i in range(1, sh.max_row + 1):
            cell_obj = sh.cell(row=i, column=1)
            if cell_obj.value is not None and len(cell_obj.value):
                try:
                    full_first_name = cell_obj.value.split()[0]
                    first_name = gen_first_name(full_first_name)
                except:
                    first_name = random.choice(["rohan", "vishal", "saurabh", "prakash", "angad"])

                try:
                    full_last_name = cell_obj.value.split()[1]
                    last_name = gen_last_name(full_last_name)

                except:
                    last_name = random.choice(["_x", "_12", "_69"])
                first_choice = gen_user_name_symbols(first_name, last_name)
                second_choice = gen_user_name_with_numbers(first_name, last_name)
                user_name = random.choice([first_choice, second_choice, first_name + " " + last_name])
                cell_obj.value = user_name
    print("Generating output file")
    Workbook.save("edited_name_output.xlsx")
    print("Successfully completed!")


def gen_first_name(full_first_name):
    one_letter_first_name = full_first_name[0]
    three_letter_first_name = full_first_name[0:3].ljust(3, 'x')
    first_name_with_underscore = "_" + full_first_name
    first_name = random.choice([full_first_name, one_letter_first_name, first_name_with_underscore,
                                three_letter_first_name])
    return first_name


def gen_last_name(full_last_name):
    three_letter_last_name = full_last_name[0:3].ljust(3, 'x')
    last_name_with_underscore = full_last_name + "_"
    last_name = random.choice([full_last_name, three_letter_last_name, last_name_with_underscore])
    return last_name


def gen_user_name_symbols(first_name, last_name):
    one_digit_number = random.randrange(0, 9)
    two_digit_number = random.randrange(1, 99)
    three_digit_number = '{:02d}'.format(random.randrange(1, 999))
    number = random.choice([one_digit_number, two_digit_number, three_digit_number])
    username = '{}{}{}{}'.format(first_name, random.choice('@&_'), number, last_name)
    return username


def gen_user_name_with_numbers(first_name, last_name):
    number = '{:02d}'.format(random.randrange(1, 999))
    user_name_array = [number, first_name, last_name]
    random.shuffle(user_name_array)
    username = '{}{}{}'.format(user_name_array[0], user_name_array[1], user_name_array[2])
    return username
    

if __name__ == '__main__':
    edit_excel()

